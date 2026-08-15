from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterator, List


def _trim_trailing_empty(values: List[str]) -> List[str]:
    while values and values[-1] == "":
        values.pop()
    return values


def iter_sheet_rows(path: Path) -> Iterator[List[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.reader(handle):
                yield _trim_trailing_empty(list(row))
        return

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required to import .xlsx calendar files"
            ) from exc

        workbook = load_workbook(path, data_only=True, read_only=True)
        worksheet = workbook.worksheets[0]
        for row in worksheet.iter_rows(values_only=True):
            yield _trim_trailing_empty(
                ["" if value is None else str(value) for value in row]
            )
        return

    raise ValueError(f"Unsupported calendar input: {path.suffix}")
